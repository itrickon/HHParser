import os
import re
import asyncio
import random
from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook


class HHVacancyCollector:
    """Сбор URL вакансий со страницы поиска HH.ru"""
    
    def __init__(self, search_url: str, max_vacancies: int = 50):
        self.search_url = search_url
        self.max_vacancies = max_vacancies
        self.vacancies = []
        self.data_saving = "hh_parse_results/hh_url_search_results.xlsx"
        self.warning_message()

    def _extract_vacancy_id(self, url: str) -> str:
        """Извлекает ID вакансии из URL hh.ru"""
        match = re.search(r"vacancy/(\d+)", url)
        return match.group(1) if match else None

    async def _get_links(self):
        """Получение ссылок на вакансии с текущей страницы"""
        # Селектор для ссылок на вакансии hh.ru
        link_selector = 'a[data-qa="serp-item__title"]'
        found_links = await self.page.query_selector_all(link_selector)

        links = []
        for link in found_links:
            href = await link.get_attribute("href")
            if href:
                # Пропускаем рекламные ссылки adsrv.hh.ru
                if "adsrv.hh.ru" in href or "adsrv.hh.ru" in href:
                    continue
                    
                vacancy_id = self._extract_vacancy_id(href)
                # Очищаем URL от параметров
                clean_url = href.split("?")[0]
                full_url = f"https://hh.ru{clean_url}" if clean_url.startswith("/") else clean_url
                links.append({"url": full_url, "id": vacancy_id})

        return links

    async def _go_to_next_page(self):
        """Переход на следующую страницу результатов поиска"""
        try:
            next_button = await self.page.query_selector('a[data-qa="pager-next"]')
            if next_button and await next_button.is_visible():
                await next_button.click()
                await asyncio.sleep(random.uniform(1.5, 2.5))
                return True
            return False
        except Exception as e:
            print(f"Ошибка при переходе на следующую страницу: {e}")
            return False

    def _create_xlsx(self):
        """Создание XLSX файла с заголовками"""
        os.makedirs("hh_parse_results", exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "HH Vacancies URLs"

        headers = ["Ссылка на вакансию"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        wb.save(self.data_saving)
        print(f"Создан файл: {self.data_saving}")

    def _save_to_xlsx(self):
        """Сохранение данных в XLSX файл"""
        if not os.path.exists(self.data_saving):
            self._create_xlsx()

        wb = load_workbook(self.data_saving)
        ws = wb.active

        start_row = ws.max_row + 1 if ws.max_row > 1 else 2

        for vacancy in self.vacancies:
            ws.cell(row=start_row, column=1, value=vacancy["url"])
            start_row += 1

        wb.save(self.data_saving)
        print(f"Данные сохранены в файл: {self.data_saving}")

    def warning_message(self):
        print("\n" + "=" * 50)
        print("EDUCATIONAL USE ONLY - NO WARRANTY PROVIDED")
        print("This parser may violate Terms of Service.")
        print("Use only for learning web scraping techniques.")
        print("Author not responsible for any legal consequences.")
        print("=" * 50 + "\n")

    async def parse_main(self, update_callback=None):
        """Сбор вакансий со страницы поиска"""
        self._create_xlsx()

        if update_callback:
            update_callback("Начало сбора вакансий...")

        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(headless=False)
            self.context = await browser.new_context()
            self.page = await self.context.new_page()

            if update_callback:
                update_callback(f"Переход на страницу: {self.search_url}")

            await self.page.goto(
                self.search_url,
                wait_until="domcontentloaded",
            )

            # Ждем появления результатов поиска
            try:
                await self.page.wait_for_selector(
                    'a[data-qa="serp-item__title"]',
                    timeout=30000
                )
                if update_callback:
                    update_callback("Страница поиска загружена")
            except Exception as e:
                print(f"Ошибка: не удалось загрузить результаты поиска: {e}")
                if update_callback:
                    update_callback(f"Ошибка загрузки страницы: {e}")
                await browser.close()
                return

            # Собираем ссылки с нескольких страниц
            while len(self.vacancies) < self.max_vacancies:
                page_links = await self._get_links()

                for link in page_links:
                    if len(self.vacancies) < self.max_vacancies:
                        self.vacancies.append(link)
                        await asyncio.sleep(0.1)

                print(f"Всего собрано ссылок: {len(self.vacancies)} из {self.max_vacancies}")
                if update_callback:
                    update_callback(f"Найдено вакансий: {len(self.vacancies)} из {self.max_vacancies}")

                if len(self.vacancies) >= self.max_vacancies:
                    if update_callback:
                        update_callback(f"Достигнуто необходимое количество вакансий: {self.max_vacancies}")
                    break

                if not await self._go_to_next_page():
                    if update_callback:
                        update_callback("Больше нет страниц для парсинга")
                    break

                await asyncio.sleep(random.uniform(1.5, 2.5))

            self._save_to_xlsx()

            print(f"Количество уникальных вакансий: {len(self.vacancies)}")
            if update_callback:
                update_callback(f"Сбор завершен. Найдено вакансий: {len(self.vacancies)}")

            await browser.close()

        return self.vacancies


async def main():
    collector = HHVacancyCollector(
        search_url="https://saratov.hh.ru/search/vacancy?area=1234",
        max_vacancies=20
    )
    await collector.collect()


if __name__ == "__main__":
    asyncio.run(main())
