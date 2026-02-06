import os
import re
import random
import asyncio
import openpyxl
import pandas as pd
from pathlib import Path
from playwright.async_api import (
    async_playwright,
    Page as AsyncPage,
    Error as PWError,
    TimeoutError as PWTimeoutError,
)

class HHParse:
    def __init__(self, input_file: str, max_num_firm: int, gui_works: bool):
        
        self.gui_works = gui_works
        self.enter_event = asyncio.Event() if gui_works else None
        self.use_gui_input = gui_works
        
        self.input_file = Path(input_file)
        self.max_num_firm = max_num_firm
        self.data_saving = "hh_parse_results/data.xlsx"
        self.list_of_companies = []
        self.start_row = 2
        self.count_page = 0
    
        # БАЗОВЫЕ ТАЙМАУТЫ
        self.CLICK_DELAY = 1.5       # Базовая задержка в секундах перед ожиданием появления номера телефона
        self.NAV_TIMEOUT = 35_000  # Таймаут загрузки страницы, мс (35 секунд)
        self.CONCURRENCY = 3   # Количество одновременно открытых вкладок браузера (2–3 оптимально)
        self.BATCH_CONCURRENCY_JITTER = True          # Иногда работаем 2 вкладками вместо 3 для естественности
        self.NAV_STAGGER_BETWEEN_TABS = (0.45, 1.0, )     # Пауза перед открытием КАЖДОЙ вкладки (чтобы не стартовали все разом)
        self.POST_NAV_IDLE = (0.35, 0.7,)                 # Небольшая «заминка» после загрузки страницы перед действиями
        self.PAGE_DELAY_BETWEEN_BATCHES = (1.2, 2.4, )    # Пауза между партиями ссылок (раньше была (2.0, 4.0))
        self.CLOSE_STAGGER_BETWEEN_TABS = (0.25, 0.55, )  # Вкладки закрываем с небольшой случайной паузой
        
        if os.path.exists(self.data_saving):
            os.remove(self.data_saving) 
            
        # ВХОДНОЙ ФАЙЛ С ССЫЛКАМИ
        self.INPUT_SHEET = None  # Имя листа в Excel; None = использовать все листы
        self.URL_COLUMN = None   # Имя колонки со ссылками; None = искать ссылки во всех колонках
        
        # ЧЕЛОВЕЧНОСТЬ / АНТИБАН-ПОВЕДЕНИЕ
        self.HUMAN = {
            "pre_page_warmup_scrolls": (1, 3, ),      # Сколько раз «прогрелись» скроллом после открытия страницы
            "scroll_step_px": (250, 900),             # Диапазон шага скролла в пикселях
            "scroll_pause_s": (0.18, 0.75),           # Пауза между скроллами
            "hover_pause_s": (0.14, 0.42),            # Пауза при наведении на элементы
            "pre_click_pause_s": (0.10, 0.28),        # Короткая пауза перед кликом
            "post_click_pause_s": (0.12, 0.32),       # Пауза сразу после клика
            "mouse_wiggle_px": (4, 12),               # Амплитуда «подёргивания» мыши
            "mouse_wiggle_steps": (2, 5),             # Сколько шагов «подёргиваний» мыши
            "between_actions_pause": (0.10, 0.30, ),  # Пауза между действиями (скролл, клик, наведение)
            "click_delay_jitter": (
                self.CLICK_DELAY * 0.9,
                self.CLICK_DELAY * 1.25
            ),  # Случайная задержка после клика по телефону (min и max)
        }   
     
    async def human_sleep(self, a: float, b: float):
        '''
        Приостанавливает выполнение на случайное количество секунд в диапазоне [a, b].
        Используется для имитации человеческих пауз и предотвращения блокировок!
        '''
        await asyncio.sleep(random.uniform(a, b))
    
    async def human_scroll_jitter(self, page: AsyncPage, count: int | None = None):
        '''
        Имитирует человеческий скроллинг страницы.
        Выполняет случайное количество скроллов со случайным шагом и направлением.
        page: Playwright Page объект
        count: Количество скроллов
        '''
        if count is None:
            count = random.randint(*self.HUMAN["pre_page_warmup_scrolls"]) # Случайное количество скролов
        try:
            height = await page.evaluate("() => document.body.scrollHeight") or 3000
            for _ in range(count):
                step = random.randint(*self.HUMAN["scroll_step_px"])
                direction = 1 if random.random() > 0.25 else -1
                y = max(0, min(height, await page.evaluate("() => window.scrollY") + step * direction))
                await page.evaluate("y => window.scrollTo({top: y, behavior: 'smooth'})", y)  # Плавный скролл через JavaScript
                await self.human_sleep(*self.HUMAN["scroll_pause_s"])
        except Exception:
            pass

     
    async def press_and_rel(self):
        """Ожидает нажатия Enter из GUI или консоли"""
        if self.gui_works:
            # Ждем, пока GUI пошлет событие
            print("Ожидаю нажатия Enter из GUI...")
            await self.wait_for_gui_enter()
        else:
            # Старый способ - ждем из консоли
            loop = asyncio.get_event_loop()
            await loop.run_in_executor(None, input, "Готов? Нажми Enter в консоли: ")
     
    async def wait_for_gui_enter(self):
        """Асинхронно ждет события от GUI"""
        while not self.enter_event.is_set():
            await asyncio.sleep(0.1)
        self.enter_event.clear()  # Сбрасываем для следующего использования
    
    def trigger_enter_from_gui(self):
        """Вызывается из GUI для имитации нажатия Enter"""
        if self.gui_works and hasattr(self, 'enter_event'):
            self.enter_event.set()
       
    def read_urls_from_excel_or_csv(self, sheet=None, url_column=None) -> list[str]:
        '''
        Читает URL вакансий с hh.ru из Excel или CSV файла.
        Args:
            sheet: Имя листа Excel (None для всех листов)
            url_column: Имя колонки с URL (None для поиска во всех колонках)
        Return: Список URL вакансий hh.ru
        '''
        if not self.input_file.exists():
            raise FileNotFoundError(f"Файл не найден: {self.input_file}")
        
        # Регулярное выражение для поиска URL hh.ru
        url_re = re.compile(r'https?://(?:[a-z]+\.)?hh\.ru/vacancy/\d+')
        urls: list[str] = []

        if self.input_file.suffix.lower() in {".xlsx", ".xls"}:
            xls = pd.ExcelFile(self.input_file)
            sheets = [sheet] if sheet is not None else xls.sheet_names
            for sh in sheets:
                df = xls.parse(sh, dtype=str)
                if url_column and url_column in df.columns:
                    col = df[url_column].dropna().astype(str)
                    urls.extend(col.tolist())
                else:
                    for col in df.columns:
                        s = df[col].dropna().astype(str)
                        for val in s:
                            urls.extend(url_re.findall(val))
        elif self.input_file.suffix.lower() in {".csv", ".txt"}:
            df = pd.read_csv(self.input_file, dtype=str, sep=None, engine="python")
            if url_column and url_column in df.columns:
                col = df[url_column].dropna().astype(str)
                urls.extend(col.tolist())
            else:
                for col in df.columns:
                    s = df[col].dropna().astype(str)
                    for val in s:
                        urls.extend(url_re.findall(val))
        else:
            raise ValueError(f"Неподдерживаемый формат файла: {self.input_file.suffix}")
        
        print(f"Прочитано {len(urls)} URL из файла: {self.input_file.name}")
        return urls 
    
    async def data_output_to_xlsx(self, get_firm_data):
        """Выводим данные в файл xlsx"""
        try:
            # Создаем или открываем файл
            if not os.path.exists(self.data_saving):
                # Создаем новую книгу
                self.wb = openpyxl.Workbook()
                self.ws = self.wb.active
                # Записываем заголовки
                headers = ['URL', 'Название вакансии', 'Название компании', 'Телефон', 'ФИО']
                for col, header in enumerate(headers, start=1):
                    self.ws.cell(row=1, column=col, value=header)
                self.start_row = 2
            else:
                # Открыть существующий файл
                self.wb = openpyxl.load_workbook(self.data_saving)
                self.ws = self.wb.active
                # Находим последнюю заполненную строку
                self.start_row = self.ws.max_row + 1
            
            # Цикл по данным фирм
            for firm_data in get_firm_data:
                # Проверяем, что firm_data не пустой
                if firm_data:
                    # Запись каждой строки в Excel
                    for col, value in enumerate(firm_data, start=1):
                        self.ws.cell(row=self.start_row, column=col, value=value)
                    self.start_row += 1
            
            # Сохраняем файл
            self.wb.save(self.data_saving)
        except Exception as e:
            print(f"Ошибка при сохранении в Excel: {e}")
            
    def get_random_user_agent(self):
        """Скрываем автоматизацию с помощью захода с разных систем"""
        user_agents = [
            # Windows Chrome - разные версии
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.6167.85 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 11_7_10) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        ]
        return random.choice(user_agents)
    
    async def process_urls_with_pool(self, context, urls: list[str], update_callback=None):
        '''
        Обрабатывает список URL с использованием пула страниц.
        Args:
            context: Контекст браузера Playwright
            urls: Список URL для обработки
            pending_queue: Список для добавления отложенных URL
        '''
        if not urls:
            return

        # Пул создаём максимального размера; часть вкладок можем не использовать
        pages = [await context.new_page() for _ in range(self.CONCURRENCY)]
        try:
            it = iter(urls)  # Итератор по URL
            while True:
                # Иногда делаем партию меньше максимума, чтобы поведение было менее ровным
                batch_size = (
                    random.randint(max(1, self.CONCURRENCY - 1), self.CONCURRENCY)
                    if self.BATCH_CONCURRENCY_JITTER
                    else self.CONCURRENCY
                )
                batch_pages = pages[:batch_size]

                batch = []  # Инициализация списка для текущей партии
                for p in batch_pages:  # Цикл по страницам партии
                    try:
                        url = next(it)
                    except StopIteration:
                        return
                    batch.append((url, p))

                    # Не открываем все вкладки синхронно — ставим паузу перед каждым goto
                    await self.human_sleep(*self.NAV_STAGGER_BETWEEN_TABS)
                    try:
                        await p.goto(url, wait_until="domcontentloaded", timeout=self.NAV_TIMEOUT)
                    except PWTimeoutError:
                        print(f"Таймаут: {url}")
                        continue

                    # Лёгкая «заминка» после навигации + пара скроллов(скрыто)
                    await self.human_sleep(*self.POST_NAV_IDLE)
                    # await self.human_scroll_jitter(p, count=random.randint(1, 2))
                
                # Получаем данные фирмы для каждого URL в партии
                for url, p in batch:
                    await self.human_sleep(*self.HUMAN["between_actions_pause"])
                    
                    try:
                        # Извлекаем данные фирмы
                        firm_data = await self.__get_firm_data_from_page(p, url)
                        
                        # Сохраняем результат
                        if firm_data:
                            # Используем await, так как метод теперь async
                            await self.data_output_to_xlsx([firm_data])  # Передаем как список с одним элементом
                            print(f"Данные фирмы: {url} -> {firm_data}")
                            if update_callback:
                                update_callback(f"Успешно: {url}")
                        else:
                            print(f"Не удалось получить данные фирмы: {url}")
                            if update_callback:
                                update_callback(f"Ошибка данных: {url}")
                                
                    except Exception as e:
                        print(f"Ошибка при обработке {url}: {e}")
                        if update_callback:
                            update_callback(f"Ошибка: {url}")


                await self.human_sleep(*self.PAGE_DELAY_BETWEEN_BATCHES)  # Пауза между партиями
        finally:
            for p in pages:
                try:
                    await self.human_sleep(*self.CLOSE_STAGGER_BETWEEN_TABS)
                    await p.close()  # Закрытие страницы
                except Exception:
                    pass


    async def __get_firm_data_from_page(self, page, url: str):
        """Извлекает данные фирмы с открытой страницы"""
        
        url = url[url.find('hh'):] # Берем ссылке, начиная с hh
        firm_data = {
            "url": url,
            "firm_vacancy": "Не найдено",
            "company_name": "Не найдено",
            "true_phone": "Телефон не найден",
            "fio": "Не указано",
        }
        
        try:
            # Извлечение названия вакансии
            vacancy_element = await page.query_selector('[data-qa="vacancy-title"] span')
            if vacancy_element:
                vacancy_text = await vacancy_element.text_content()
                if vacancy_text:
                    firm_data["firm_vacancy"] = vacancy_text.strip()
            
            # Извлечение названия компании
            company_element = await page.query_selector('[data-qa="vacancy-company-name"] span')
            if company_element:
                company_text = await company_element.text_content()
                if company_text:
                    firm_data["company_name"] = ' '.join(company_text.strip().split('\xa0'))

            # Номер телефона и ФИО - ищем через кнопку "Связаться"
            try:
                # Ищем кнопку "Связаться"
                contact_button = await page.query_selector('button[data-qa*="show-employer-contacts"]')
                if contact_button:
                    # Пытаемся кликнуть, если кнопка есть
                    await contact_button.click()
                    await self.human_sleep(0.5, 1.0)  # Ждем появления контактной информации
                    
                    # После клика ищем ФИО
                    fio_element = await page.query_selector('[data-qa="vacancy-contacts__fio"]')
                    if fio_element:
                        fio_text = await fio_element.text_content()
                        if fio_text and fio_text.strip():
                            firm_data["fio"] = fio_text.strip()
                    
                    # После клика ищем телефон
                    phone_element = await page.query_selector('[data-qa="vacancy-contacts__phone-number"]')
                    if phone_element:
                        phone_text = await phone_element.text_content()
                        firm_data["true_phone"] = phone_text
       
            except Exception as e:
                print(f"Ошибка при поиске телефона или ФИО: {e}")

        except Exception as e:
            print(f"Ошибка при получении данных фирмы: {e}")
        
        # Возвращаем в формате списка (убрали сайт, добавили ФИО)
        return [
            firm_data["url"],
            firm_data["firm_vacancy"],
            firm_data["company_name"],
            firm_data["true_phone"],
            firm_data["fio"],
        ]
    
    async def parse_main(self, update_callback=None):  
        """Парсинг сайта"""
        urls = self.read_urls_from_excel_or_csv(self.INPUT_SHEET, self.URL_COLUMN)
        urls = urls[:self.max_num_firm]
        
        print(f"Новых ссылок к обработке: {len(urls)};")
        if update_callback:
            update_callback(f"Новых ссылок к обработке: {len(urls)};")
        # atexit.register(self.flush_progress)  # Регистрация функции при завершении программы
        print(urls)
        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(
                headless=False,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-features=IsolateOrigins,site-per-process',
                    '--disable-web-security',
                    '--disable-site-isolation-trials',
                ]
            )  # headless=True - без графического итерфейса
            try:
                vp_w = random.randint(1200, 1400)
                vp_h = random.randint(760, 900)
                context = await browser.new_context(
                    viewport={"width": vp_w, "height": vp_h},
                    user_agent=self.get_random_user_agent(),
                    locale="ru-RU",
                    timezone_id="Europe/Moscow",
                    extra_http_headers={'Cache-Control': 'no-cache'},
                )
                
                # Ручной логин на первой ссылке (если есть что открывать)
                seed_url = urls[0] if urls else None
                if seed_url:
                    
                    page = await context.new_page() # Создание новой страницы
                    
                    await asyncio.sleep(random.uniform(0.4, 0.8))
                    try:
                        await asyncio.sleep(random.uniform(0.5, 0.8))
                        
                        # Потом на объявление
                        await page.goto(seed_url, 
                                    wait_until="domcontentloaded", 
                                    timeout=self.NAV_TIMEOUT)
                    except PWTimeoutError:
                        try:
                            await page.goto(seed_url, 
                                        wait_until="domcontentloaded", 
                                        timeout=self.NAV_TIMEOUT)
                        except PWTimeoutError:
                            print(f"Таймаут при загрузке {seed_url}")
                    
                    await self.human_sleep(0.4, 0.7)
                    
                    print("\nТвои действия:")  # Инструкция пользователю
                    print(" • если есть капча — реши;")
                    print(" • залогинься в Авито;")
                    print(" • оставь открытую страницу объявления.")
                    
                    # Здесь ждем подтверждения входа
                    if self.gui_works:
                        if update_callback:
                            update_callback("Ожидание подтверждения входа... Нажмите 'Вход выполнен'")
                        await self.press_and_rel()  # Ждем нажатия кнопки в GUI
                    else:
                        # Старый способ для консоли
                        loop = asyncio.get_event_loop()
                        await loop.run_in_executor(None, input, "Готов? Нажми Enter в консоли: ")
                    
                    try:
                        await page.close()
                    except Exception:
                        pass
                        # Основной список из Excel
                try:
                    await self.process_urls_with_pool(context, urls, update_callback)
                except Exception as e:
                    print(f"Ошибка {e}")
                    
            finally:
                await browser.close()
                self.browser = None
        
async def main(): 
    parser = HHParse(
        input_file="abc.xlsx",
        max_num_firm=20,  # Сколько вакансий собрать
        gui_works = False,
    )
    await parser.parse_main()


if __name__ == "__main__":
    asyncio.run(main())