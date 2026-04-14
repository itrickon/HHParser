"""
Модуль для работы с Excel файлами
Чтение, запись, обновление данных
"""

import logging
from pathlib import Path
from typing import List, Dict, Optional, Any

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import config
from logger import get_logger


class ExcelManager:
    """Менеджер для работы с Excel файлами"""
    
    def __init__(self):
        self.logger = get_logger()

    def read_urls_from_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        url_column: Optional[str] = None,
    ) -> List[str]:
        """
        Читает URL вакансий из Excel файла
        
        Args:
            file_path: Путь к Excel файлу
            sheet_name: Имя листа (None = все листы)
            url_column: Имя колонки с URL (None = поиск во всех колонках)
            
        Returns:
            Список URL вакансий hh.ru
        """
        import re
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"Файл не найден: {file_path}")
        
        url_pattern = re.compile(config.HH_VACANCY_URL_PATTERN)
        urls: List[str] = []
        
        try:
            xls = pd.ExcelFile(file_path)
            sheets = [sheet_name] if sheet_name else xls.sheet_names
            
            for sheet in sheets:
                df = xls.parse(sheet, dtype=str)
                
                if url_column and url_column in df.columns:
                    # Ищем в конкретной колонке
                    col_data = df[url_column].dropna().astype(str)
                    for val in col_data:
                        found_urls = url_pattern.findall(val)
                        urls.extend(found_urls)
                else:
                    # Ищем во всех колонках
                    for col in df.columns:
                        col_data = df[col].dropna().astype(str)
                        for val in col_data:
                            found_urls = url_pattern.findall(val)
                            urls.extend(found_urls)
            
            # Убираем дубликаты сохраняя порядок
            unique_urls = list(dict.fromkeys(urls))
            self.logger.info(f"Найдено {len(unique_urls)} уникальных URL из {len(urls)}")
            
            return unique_urls
            
        except Exception as e:
            self.logger.error(f"Ошибка при чтении Excel файла: {e}")
            raise

    def save_to_excel(
        self,
        data: List[Dict[str, Any]],
        output_file: Optional[str] = None,
        sheet_name: str = config.EXCEL_SHEET_NAME,
    ) -> str:
        """
        Сохраняет данные в Excel файл с красивым форматированием
        
        Args:
            data: Список словарей с данными
            output_file: Путь к файлу (по умолчанию config.OUTPUT_FILE)
            sheet_name: Имя листа
            
        Returns:
            Путь к сохраненному файлу
        """
        if not data:
            self.logger.warning("Нет данных для сохранения")
            return ""
        
        output_file = output_file or config.OUTPUT_FILE
        output_path = Path(output_file)
        
        # Создаем директорию если её нет
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            # Создаем DataFrame
            df = pd.DataFrame(data)
            
            # Переупорядочиваем колонки если нужно
            if all(col in df.columns for col in config.EXCEL_COLUMNS):
                df = df[config.EXCEL_COLUMNS]
            
            # Сохраняем в Excel
            df.to_excel(output_file, index=False, sheet_name=sheet_name)
            
            # Применяем форматирование
            self._format_excel_file(output_file, sheet_name)
            
            self.logger.info(f"Данные сохранены в {output_file} ({len(data)} записей)")
            
            return str(output_file)
            
        except Exception as e:
            self.logger.error(f"Ошибка при сохранении в Excel: {e}")
            raise

    def append_to_excel(
        self,
        data: List[Dict[str, Any]],
        output_file: Optional[str] = None,
        sheet_name: str = config.EXCEL_SHEET_NAME,
    ) -> str:
        """
        Добавляет данные в существующий Excel файл
        
        Args:
            data: Список словарей с данными
            output_file: Путь к файлу
            sheet_name: Имя листа
            
        Returns:
            Путь к обновленному файлу
        """
        if not data:
            self.logger.warning("Нет данных для добавления")
            return ""
        
        output_file = output_file or config.OUTPUT_FILE
        output_path = Path(output_file)
        
        try:
            if output_path.exists():
                # Читаем существующий файл
                existing_df = pd.read_excel(output_file, sheet_name=sheet_name)
                new_df = pd.DataFrame(data)
                
                # Объединяем
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                
                # Сохраняем
                combined_df.to_excel(output_file, index=False, sheet_name=sheet_name)
            else:
                # Создаем новый файл
                return self.save_to_excel(data, output_file, sheet_name)
            
            self.logger.info(f"Добавлено {len(data)} записей в {output_file}")
            
            return str(output_file)
            
        except Exception as e:
            self.logger.error(f"Ошибка при добавлении в Excel: {e}")
            raise

    @staticmethod
    def _format_excel_file(file_path: str, sheet_name: str = config.EXCEL_SHEET_NAME):
        """
        Применяет форматирование к Excel файлу
        
        Args:
            file_path: Путь к файлу
            sheet_name: Имя листа
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb[sheet_name]
            
            # Стили
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_alignment = Alignment(vertical="top", wrap_text=True)
            
            # Форматируем заголовки
            for col_idx, cell in enumerate(ws[1], 1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Форматируем данные
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = cell_alignment
            
            # Автоматическая ширина колонок
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in ws[worksheet_dimension[0] for worksheet_dimension in ws.dimensions.split(':')][0]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Устанавливаем ширину (минимум 10, максимум 50)
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Автофильтр
            ws.auto_filter.ref = ws.dimensions
            
            # Закрепляем первую строку
            ws.freeze_panes = "A2"
            
            wb.save(file_path)
            wb.close()
            
        except Exception as e:
            logging.debug(f"Ошибка при форматировании Excel: {e}")

    def get_excel_info(self, file_path: str) -> Dict[str, Any]:
        """
        Получает информацию о Excel файле
        
        Args:
            file_path: Путь к файлу
            
        Returns:
            Словарь с информацией {sheets, row_count, columns}
        """
        try:
            xls = pd.ExcelFile(file_path)
            
            info = {
                "sheets": xls.sheet_names,
                "total_rows": 0,
                "columns": [],
            }
            
            for sheet in xls.sheet_names:
                df = xls.parse(sheet)
                info["total_rows"] += len(df)
                if not info["columns"]:
                    info["columns"] = list(df.columns)
            
            return info
            
        except Exception as e:
            self.logger.error(f"Ошибка при получении информации о файле: {e}")
            return {}
