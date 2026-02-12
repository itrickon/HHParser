import os
import re
import random
import asyncio
import openpyxl
from typing import List
from openpyxl import Workbook
from playwright.async_api import (
    async_playwright,
    Page as AsyncPage,
)


class HHParse:
    def __init__(self, link: str, max_num_firm: int):
        self.link = link
        self.max_num_firm = max_num_firm
        self.data_saving = "hh_parse_results/data.xlsx"
        self.list_of_companies = []
        self.start_row = 2
        self.count_page = 0

        if os.path.exists(self.data_saving):
            os.remove(self.data_saving)

        # Конфигурация для естественного поведения
        self.PAGE_DELAY_BETWEEN_BATCHES = (1.2, 2.4)
        self.NAV_STAGGER_BETWEEN_TABS = (0.45, 1.0)
        self.POST_NAV_IDLE = (0.35, 0.7)
        self.CLOSE_STAGGER_BETWEEN_TABS = (0.25, 0.55)
        self.CLICK_DELAY = 1.5
        self.NAV_TIMEOUT = 35000

        # Человеческие параметры
        self.HUMAN = {
            "pre_page_warmup_scrolls": (1, 3),
            "scroll_step_px": (250, 900),
            "scroll_pause_s": (0.18, 0.75),
            "hover_pause_s": (0.14, 0.42),
            "pre_click_pause_s": (0.10, 0.28),
            "post_click_pause_s": (0.12, 0.32),
            "mouse_wiggle_px": (4, 12),
            "mouse_wiggle_steps": (2, 5),
            "between_actions_pause": (0.10, 0.30),
            "click_delay_jitter": (self.CLICK_DELAY * 0.9, self.CLICK_DELAY * 1.25),
        }

    async def human_sleep(self, a: float, b: float):
        """Случайная задержка между действиями"""
        await asyncio.sleep(random.uniform(a, b))

    async def human_scroll_jitter(self, page: AsyncPage, count: int = None):
        """Имитация человеческого скроллинга"""
        if count is None:
            count = random.randint(*self.HUMAN["pre_page_warmup_scrolls"])

        try:
            height = await page.evaluate("() => document.body.scrollHeight") or 3000
            for _ in range(count):
                step = random.randint(*self.HUMAN["scroll_step_px"])
                direction = 1 if random.random() > 0.25 else -1
                y = max(
                    0,
                    min(
                        height,
                        await page.evaluate("() => window.scrollY") + step * direction,
                    ),
                )
                await page.evaluate(
                    "y => window.scrollTo({top: y, behavior: 'smooth'})", y
                )
                await self.human_sleep(*self.HUMAN["scroll_pause_s"])
        except Exception:
            pass

    async def human_hover(self, page: AsyncPage, el):
        """Имитация человеческого наведения мыши"""
        try:
            box = await el.bounding_box()
            if not box:
                return

            cx = box["x"] + box["width"] * random.uniform(0.35, 0.65)
            cy = box["y"] + box["height"] * random.uniform(0.35, 0.65)
            await page.mouse.move(cx, cy)
            await self.human_sleep(*self.HUMAN["hover_pause_s"])
        except Exception:
            pass

    def get_random_user_agent(self):
        """Генерация случайного User-Agent"""
        user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.6167.85 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        ]
        return random.choice(user_agents)

    async def check_xlsx(self):
        """Создание Excel файла с заголовками"""
        os.makedirs("hh_parse_results", exist_ok=True)

        self.wb = Workbook()
        self.ws = self.wb.active

        headers = ["Вакансия", "Компания", "Город", "Номер"]

        for col, header in enumerate(headers, start=1):
            self.ws.cell(row=1, column=col, value=header)

        self.wb.save(self.data_saving)

    async def data_output_to_xlsx(self, firm_data_list):
        """Запись данных в Excel"""
        if os.path.exists(self.data_saving):
            self.wb = openpyxl.load_workbook(self.data_saving)
            self.ws = self.wb.active
        else:
            await self.check_xlsx()

        for firm_data in firm_data_list:
            for col, value in enumerate(firm_data, start=1):
                self.ws.cell(row=self.start_row, column=col, value=value)
            self.start_row += 1

        self.wb.save(self.data_saving)
        print(f"Записано {len(firm_data_list)} вакансий в файл")

    async def extract_phone_from_contact_popup(self, page: AsyncPage) -> str:
        """Извлечение телефона из всплывающего окна контактов"""
        try:
            await self.human_sleep(1.0, 1.5)

            # Основные селекторы для телефона
            phone_selectors = [
                '[data-qa="vacancy-contacts__phone-number"]',
                'span[data-qa="vacancy-contacts__phone-number"]',
                '.magritte-text:has-text("Основной телефон") + .magritte-v-spacing + .magritte-text',
                'div.magritte-card[style*="border-radius: 24px"] .magritte-text:has-text("+7")',
            ]

            for selector in phone_selectors:
                try:
                    phone_element = await page.query_selector(selector)
                    if phone_element and await phone_element.is_visible():
                        phone_text = (await phone_element.text_content()).strip()
                        if phone_text:
                            # Очищаем номер
                            phone_digits = re.sub(r"\D", "", phone_text)
                            if len(phone_digits) >= 10:
                                # Форматируем номер
                                if (
                                    phone_digits.startswith("7")
                                    and len(phone_digits) == 11
                                ):
                                    return f"+{phone_digits[0]} {phone_digits[1:4]} {phone_digits[4:7]}-{phone_digits[7:9]}-{phone_digits[9:]}"
                                elif (
                                    phone_digits.startswith("8")
                                    and len(phone_digits) == 11
                                ):
                                    return f"+7 {phone_digits[1:4]} {phone_digits[4:7]}-{phone_digits[7:9]}-{phone_digits[9:]}"
                                elif len(phone_digits) == 10:
                                    return f"+7 {phone_digits[0:3]} {phone_digits[3:6]}-{phone_digits[6:8]}-{phone_digits[8:]}"
                                return phone_text
                except:
                    continue

            return "Номер не найден"

        except Exception as e:
            print(f"Ошибка при извлечении телефона: {e}")
            return "Ошибка"

    async def click_contact_button(self, page: AsyncPage) -> bool:
        """Клик по кнопке 'Связаться'"""
        try:
            contact_selectors = [
                'button[data-qa="vacancy-serp__vacancy_contacts"]',
                'a[data-qa="vacancy-contacts-button"]',
                'button:has-text("Связаться")',
                'a:has-text("Связаться")',
            ]

            for selector in contact_selectors:
                try:
                    contact_button = await page.query_selector(selector)
                    if contact_button and await contact_button.is_visible():
                        await self.human_hover(page, contact_button)
                        await self.human_sleep(0.2, 0.4)
                        await contact_button.click()
                        await self.human_sleep(0.5, 0.8)

                        # Ждем появления попапа
                        await page.wait_for_selector(
                            'div[class*="magritte-drop-container"], '
                            '[data-qa="vacancy-contacts__phone"]',
                            timeout=5000,
                        )
                        return True
                except:
                    continue

            return False

        except Exception as e:
            print(f"Ошибка при клике на кнопку связи: {e}")
            return False

    async def close_contact_popup(self, page: AsyncPage):
        """Закрытие всплывающего окна"""
        try:
            await page.keyboard.press("Escape")
            await self.human_sleep(0.3, 0.5)
        except:
            pass

    async def extract_city_from_location(self, location_text: str) -> str:
        """Извлечение города из текста локации"""
        if not location_text:
            return "Не указан"

        location_text = location_text.strip()

        # Убираем лишние слова
        remove_words = [
            "район",
            "область",
            "край",
            "республика",
            "г.",
            "м.",
            "ул.",
            "пр.",
            "д.",
        ]
        for word in remove_words:
            location_text = location_text.replace(word, "").replace(word.title(), "")

        # Разделяем по разделителям
        parts = re.split(r"[,;.\-()]", location_text)
        if parts:
            return parts[0].strip()

        return location_text

    async def parse_vacancy_card(self, vacancy_card) -> dict:
        """Парсинг данных из карточки вакансии без перехода на отдельную страницу"""
        page_data = {"vacancy": "", "company": "", "city": "", "phone": ""}

        try:
            # Извлекаем данные из карточки
            # Название вакансии
            try:
                title_el = await vacancy_card.query_selector(
                    '[data-qa="serp-item__title"]'
                )
                if title_el:
                    page_data["vacancy"] = (await title_el.text_content()).strip()
            except:
                pass

            # Компания
            try:
                company_el = await vacancy_card.query_selector(
                    '[data-qa="vacancy-serp__vacancy-employer"]'
                )
                if company_el:
                    company_text = await company_el.text_content()
                    # Очищаем от лишних пробелов
                    page_data["company"] = " ".join(company_text.split()).strip()
            except:
                pass

            # Город
            try:
                city_el = await vacancy_card.query_selector(
                    '[data-qa="vacancy-serp__vacancy-address"]'
                )
                if city_el:
                    city_text = await city_el.text_content()
                    page_data["city"] = await self.extract_city_from_location(city_text)
                else:
                    page_data["city"] = "Не указан"
            except:
                page_data["city"] = "Ошибка"

            # Телефон будем получать при переходе на страницу вакансии
            page_data["phone"] = "Требуется переход"

            return page_data

        except Exception as e:
            print(f"Ошибка при парсинге карточки: {e}")
            return page_data

    async def parse_vacancy_page(self, vacancy_url: str) -> dict:
        """Парсинг полной страницы вакансии для получения телефона"""
        page_data = {"vacancy": "", "company": "", "city": "", "phone": ""}

        try:
            vacancy_page = await self.context.new_page()
            await self.human_sleep(*self.NAV_STAGGER_BETWEEN_TABS)

            await vacancy_page.goto(
                vacancy_url, wait_until="domcontentloaded", timeout=self.NAV_TIMEOUT
            )
            await self.human_sleep(*self.POST_NAV_IDLE)
            await self.human_scroll_jitter(vacancy_page)

            # Название вакансии
            try:
                title_el = await vacancy_page.query_selector(
                    '[data-qa="vacancy-title"]'
                )
                if title_el:
                    page_data["vacancy"] = (await title_el.text_content()).strip()
            except:
                pass

            # Компания
            try:
                company_el = await vacancy_page.query_selector(
                    '[data-qa="vacancy-company-name"]'
                )
                if company_el:
                    company_text = await company_el.text_content()
                    page_data["company"] = " ".join(company_text.split()).strip()
            except:
                pass

            # Город
            try:
                city_el = await vacancy_page.query_selector(
                    '[data-qa="vacancy-view-location"]'
                )
                if city_el:
                    city_text = await city_el.text_content()
                    page_data["city"] = await self.extract_city_from_location(city_text)
                else:
                    page_data["city"] = "Не указан"
            except:
                page_data["city"] = "Ошибка"

            # Телефон через кнопку "Связаться"
            try:
                if await self.click_contact_button(vacancy_page):
                    page_data["phone"] = await self.extract_phone_from_contact_popup(
                        vacancy_page
                    )
                    await self.close_contact_popup(vacancy_page)
                else:
                    page_data["phone"] = "Нет кнопки связи"
            except Exception as e:
                print(f"Ошибка при получении телефона: {e}")
                page_data["phone"] = "Ошибка"

            await self.human_sleep(*self.CLOSE_STAGGER_BETWEEN_TABS)
            await vacancy_page.close()

            return page_data

        except Exception as e:
            print(f"Ошибка при парсинге страницы вакансии {vacancy_url}: {e}")
            await vacancy_page.close()
            return page_data

    async def get_vacancy_cards(self) -> List:
        """Получение всех карточек вакансий на странице"""
        try:
            # Все объявления имеют data-qa="vacancy-serp__vacancy"
            vacancy_cards = await self.page.query_selector_all(
                '[data-qa="vacancy-serp__vacancy"]'
            )
            print(f"Найдено карточек вакансий: {len(vacancy_cards)}")
            return vacancy_cards[
                : min(20, self.max_num_firm - len(self.list_of_companies))
            ]
        except Exception as e:
            print(f"Ошибка при поиске карточек: {e}")
            return []

    async def get_vacancy_url_from_card(self, vacancy_card):
        """Получение ссылки на вакансию из карточки"""
        try:
            # Ищем ссылку в карточке
            link_el = await vacancy_card.query_selector('[data-qa="serp-item__title"]')
            if link_el:
                href = await link_el.get_attribute("href")
                if href:
                    # Убираем параметры отслеживания
                    clean_href = href.split("?")[0] if "?" in href else href
                    if clean_href.startswith("http"):
                        return clean_href
                    else:
                        return f"https://hh.ru{clean_href}"
        except Exception as e:
            print(f"Ошибка при получении ссылки: {e}")

        return None

    async def go_to_next_page(self) -> bool:
        """Переход на следующую страницу результатов"""
        try:
            # Ищем кнопку "Далее"
            next_button = await self.page.query_selector('a[data-qa="pager-next"]')

            if not next_button:
                print("Кнопка 'Далее' не найдена")
                return False

            # Проверяем активность
            href = await next_button.get_attribute("href")
            if not href:
                print("Кнопка 'Далее' неактивна")
                return False

            await self.human_hover(self.page, next_button)
            await self.human_sleep(0.2, 0.4)

            print("Переходим на следующую страницу...")
            await next_button.click()
            await self.human_sleep(2, 3)

            # Ждем загрузки
            await self.page.wait_for_load_state("domcontentloaded", timeout=10000)
            await self.human_sleep(1, 1.5)

            # Скроллим
            await self.human_scroll_jitter(self.page)

            self.count_page += 1
            print(f"Успешно перешли на страницу {self.count_page + 1}")

            return True

        except Exception as e:
            print(f"Ошибка при переходе на следующую страницу: {e}")
            return False

    async def parse_main(self, update_callback=None):
        """Основная функция парсинга"""
        async with async_playwright() as playwright:
            try:
                browser = await playwright.chromium.launch(headless=False)
                vp_w = random.randint(1200, 1400)
                vp_h = random.randint(760, 900)

                self.context = await browser.new_context(
                    viewport={"width": vp_w, "height": vp_h},
                    user_agent=self.get_random_user_agent(),
                    locale="ru-RU",
                    timezone_id="Europe/Moscow",
                    extra_http_headers={"Cache-Control": "no-cache"},
                )

                self.page = await self.context.new_page()
                await self.page.goto(
                    self.link, wait_until="domcontentloaded", timeout=self.NAV_TIMEOUT
                )

                await self.human_scroll_jitter(self.page)
                await self.check_xlsx()

                page_num = 1
                processed_count = 0

                while len(self.list_of_companies) < self.max_num_firm:
                    print(f"\nСтраница: {page_num}")
                    print(f"Собрано: {len(self.list_of_companies)}/{self.max_num_firm}")

                    # Получаем все карточки вакансий на странице
                    vacancy_cards = await self.get_vacancy_cards()

                    if not vacancy_cards:
                        print("Не найдено карточек вакансий на странице")
                        break

                    # Обрабатываем каждую карточку
                    for card in vacancy_cards:
                        if len(self.list_of_companies) >= self.max_num_firm:
                            break

                        processed_count += 1
                        print(f"\n[#{processed_count}] Парсим карточку...")

                        # Сначала получаем базовые данные из карточки
                        basic_data = await self.parse_vacancy_card(card)

                        if not basic_data["vacancy"]:
                            print("Пропускаем - нет названия вакансии")
                            continue

                        # Получаем ссылку для перехода за телефоном
                        vacancy_url = await self.get_vacancy_url_from_card(card)

                        if vacancy_url:
                            print(f"Переходим за телефоном: {vacancy_url[:80]}...")
                            full_data = await self.parse_vacancy_page(vacancy_url)

                            # Объединяем данные (предпочитаем данные со страницы)
                            if full_data["vacancy"]:
                                basic_data["vacancy"] = full_data["vacancy"]
                            if full_data["company"]:
                                basic_data["company"] = full_data["company"]
                            if full_data["city"] and full_data["city"] != "Не указан":
                                basic_data["city"] = full_data["city"]
                            if (
                                full_data["phone"]
                                and full_data["phone"] != "Требуется переход"
                            ):
                                basic_data["phone"] = full_data["phone"]

                        # Формируем финальные данные
                        firm_data = [
                            basic_data["vacancy"],
                            basic_data["company"],
                            basic_data["city"],
                            basic_data["phone"],
                        ]

                        self.list_of_companies.append(firm_data)
                        print(f"Добавлено: {basic_data['vacancy'][:50]}...")
                        print(f"  Компания: {basic_data['company'][:30]}...")
                        print(f"  Город: {basic_data['city']}")
                        print(f"  Телефон: {basic_data['phone']}")

                        # Сохраняем каждые 3 записи
                        if len(self.list_of_companies) % 3 == 0:
                            await self.data_output_to_xlsx(self.list_of_companies)
                            self.list_of_companies = []

                        await self.human_sleep(1, 2)

                    # Сохраняем остатки
                    if self.list_of_companies:
                        await self.data_output_to_xlsx(self.list_of_companies)
                        self.list_of_companies = []

                    # Проверяем лимит
                    if len(self.list_of_companies) >= self.max_num_firm:
                        print(f"Достигнут лимит в {self.max_num_firm} вакансий")
                        break

                    # Переход на следующую страницу
                    print("\nПроверяем наличие следующей страницы...")
                    if await self.go_to_next_page():
                        page_num += 1
                        await self.human_sleep(2, 3)
                    else:
                        print("Больше нет страниц для парсинга")
                        break

                    # Задержка между страницами
                    await self.human_sleep(*self.PAGE_DELAY_BETWEEN_BATCHES)

                # Финальное сохранение
                if self.list_of_companies:
                    await self.data_output_to_xlsx(self.list_of_companies)

                await browser.close()

                # Загружаем финальный файл для подсчета
                if os.path.exists(self.data_saving):
                    self.wb = openpyxl.load_workbook(self.data_saving)
                    self.ws = self.wb.active
                    total_records = self.ws.max_row - 1
                    print(f"ПАРСИНГ ЗАВЕРШЕН!")
                    print(f"Всего собрано вакансий: {total_records}")
                    print(f"Файл сохранен: {self.data_saving}")

            except Exception as e:
                error_msg = f"Произошла ошибка: {e}"
                print(error_msg)
                if update_callback:
                    update_callback(error_msg)
                raise


async def main():
    parser = HHParse(
        link="https://hh.ru/search/vacancy?text=python+разработчик&area=1",
        max_num_firm=20,  # Сколько вакансий собрать
    )
    await parser.parse_main()


if __name__ == "__main__":
    asyncio.run(main())
